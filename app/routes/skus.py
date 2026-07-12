"""CRUD SKU + batch-расчёт юнит-экономики."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from app.database import get_db
from app.deps import get_workspace_for_user
from app.models import Workspace, Sku, Category, StoreSettings
from app.schemas import SkuCreate, SkuUpdate, SkuResponse, CalcRequest
from app.services.calculator import calc_one, summarize
from app.seed import TAX_SYSTEMS

router = APIRouter(prefix="/api/workspaces/{workspace_id}/skus", tags=["skus"])


def _categories_map(db: Session, workspace_id: int):
    cats = db.query(Category).filter_by(workspace_id=workspace_id).all()
    return {c.name: {"fby_rate": c.fby_rate, "fbs_rate": c.fbs_rate} for c in cats}


def _store_settings_dict(db: Session, workspace_id: int) -> dict:
    s = db.query(StoreSettings).filter_by(workspace_id=workspace_id).first()
    if not s:
        s = StoreSettings(workspace_id=workspace_id); db.add(s); db.commit(); db.refresh(s)
    return {
        "tax_system": s.tax_system, "tax_rate": s.tax_rate,
        "acquiring_rate": s.acquiring_rate, "return_pct": s.return_pct,
        "return_cost_rub": s.return_cost_rub, "default_drr_pct": s.default_drr_pct,
    }


@router.get("", response_model=list[SkuResponse])
def list_skus(
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
    limit: int = 500, offset: int = 0,
):
    q = db.query(Sku).filter_by(workspace_id=ws.id).order_by(Sku.id)
    return q.offset(offset).limit(limit).all()


@router.post("", response_model=SkuResponse, status_code=status.HTTP_201_CREATED)
def create_sku(
    req: SkuCreate,
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    existing = db.query(Sku).filter_by(workspace_id=ws.id, sku=req.sku).first()
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, "SKU уже существует")
    s = Sku(workspace_id=ws.id, **req.model_dump())
    db.add(s); db.commit(); db.refresh(s)
    return s


@router.patch("/{sku_id}", response_model=SkuResponse)
def update_sku(
    sku_id: int, req: SkuUpdate,
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    s = db.query(Sku).filter_by(workspace_id=ws.id, id=sku_id).first()
    if not s:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "SKU не найден")
    for k, v in req.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit(); db.refresh(s)
    return s


@router.delete("/{sku_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_sku(
    sku_id: int,
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    s = db.query(Sku).filter_by(workspace_id=ws.id, id=sku_id).first()
    if not s:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Not found")
    db.delete(s); db.commit()


@router.post("/bulk-upsert")
def bulk_upsert(
    items: list[SkuCreate],
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    created = updated = 0
    for it in items:
        existing = db.query(Sku).filter_by(workspace_id=ws.id, sku=it.sku).first()
        if existing:
            data = it.model_dump()
            for k, v in data.items():
                setattr(existing, k, v)
            updated += 1
        else:
            db.add(Sku(workspace_id=ws.id, **it.model_dump()))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "total": created + updated}


@router.delete("", status_code=status.HTTP_204_NO_CONTENT)
def delete_all_skus(
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    db.query(Sku).filter_by(workspace_id=ws.id).delete()
    db.commit()


@router.post("/import-costs")
async def import_costs_xlsx(
    file: UploadFile = File(...),
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    """Импорт себестоимости из xlsx (выгрузка Ya.Market с заполненной колонкой W «Себестоимость»).
    Ищет SKU в колонке D (Ваш SKU), берёт cost из колонки W. Обновляет только те SKU, у которых
    cost > 0 в файле; уже заполненная себестоимость перезаписывается новой (если она > 0)."""
    import openpyxl, io
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Не удалось открыть xlsx: {e}")

    # Ищем лист «Список товаров», иначе первый
    ws_name = "Список товаров" if "Список товаров" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[ws_name]

    # Строим индекс существующих SKU
    all_skus = db.query(Sku).filter_by(workspace_id=ws.id).all()
    by_sku = {s.sku: s for s in all_skus}

    updated = 0
    matched = 0
    scanned = 0
    file_rows_with_cost = 0
    # Данные начинаются с 4-й строки в выгрузке Маркета (r1-r3 — шапки)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        if not row: continue
        # D = index 3, W = index 22 (нумерация с 0)
        sku_val = row[3] if len(row) > 3 else None
        cost_val = row[22] if len(row) > 22 else None
        if sku_val is None: continue
        scanned += 1
        sku_key = str(sku_val).strip()
        if not sku_key: continue
        try:
            cost = float(cost_val) if cost_val not in (None, "") else 0.0
        except (ValueError, TypeError):
            cost = 0.0
        if cost > 0: file_rows_with_cost += 1
        if sku_key in by_sku:
            matched += 1
            if cost > 0 and float(by_sku[sku_key].cost_rub or 0) != cost:
                by_sku[sku_key].cost_rub = cost
                updated += 1

    db.commit()
    return {
        "sheet_used": ws_name,
        "rows_scanned": scanned,
        "rows_with_cost_in_file": file_rows_with_cost,
        "matched_by_sku": matched,
        "updated_cost_rub": updated,
        "total_skus_in_db": len(all_skus),
    }


@router.post("/calc")
def calc_batch(
    req: CalcRequest,
    ws: Workspace = Depends(get_workspace_for_user),
    db: Session = Depends(get_db),
):
    q = db.query(Sku).filter_by(workspace_id=ws.id)
    if req.sku_ids:
        q = q.filter(Sku.id.in_(req.sku_ids))
    skus = q.all()

    store = _store_settings_dict(db, ws.id)
    cats = _categories_map(db, ws.id)

    tax_override = TAX_SYSTEMS.get(req.tax_system) if req.tax_system else None

    results = []
    for s in skus:
        d = {
            "id": s.id, "sku": s.sku, "name": s.name,
            "category": s.category, "model": s.model,
            "length_cm": s.length_cm, "width_cm": s.width_cm, "height_cm": s.height_cm,
            "weight_kg": s.weight_kg,
            "price_rub": float(s.price_rub), "cost_rub": float(s.cost_rub),
            "drr_pct": s.drr_pct,
            "stock_total": getattr(s, "stock_total", 0) or 0,
        }
        results.append(calc_one(d, store, cats,
            tax_rate_override=tax_override,
            acquiring_override=req.acquiring_rate,
            drr_override=req.drr_pct))
    return {"results": results, "summary": summarize(results)}
