"""
marja.app — итоговый P&L Ozon по SKU с себестоимостью и чистой прибылью.

Объединяет:
- import_nachisleniya: факт из отчёта Ozon (выручка + расходы по статьям);
- build_cogs: себестоимость из файла цен (матч по ID, затем по токену модели);
- расчёт: чистая прибыль = вклад до COGS − себестоимость.
"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .ozon_import import import_nachisleniya, REVENUE_LINES
from .ozon_cogs import build_cogs

SKU_COST_LINES = ["commission", "logistics", "return_logistics", "last_mile",
                  "fbs_processing", "acquiring", "returns"]


def build(report_file: str, cost_file: str, out_xlsx: str) -> dict:
    data = import_nachisleniya(report_file)
    sku_rows = [r for r in data["rows"] if r["sku"]]
    cogs = build_cogs(cost_file, [
        {"sku": r["sku"], "article": r.get("article"), "qty_sold": r["qty_sold"]}
        for r in sku_rows])

    view = []
    for r in sku_rows:
        sid = str(r["sku"]).strip()
        if sid.endswith(".0"):
            sid = sid[:-2]
        gross = round(sum(r.get(k, 0) for k in REVENUE_LINES), 2)
        costs = {k: round(r.get(k, 0), 2) for k in SKU_COST_LINES}
        contrib = round(gross + sum(costs.values()), 2)
        cg = cogs.get(sid, {})
        cogs_total = cg.get("cogs_total")
        net = round(contrib - cogs_total, 2) if cogs_total is not None else None
        view.append({
            "sku": sid, "name": r["name"], "qty": r["qty_sold"], "gross": gross,
            **costs, "contrib": contrib,
            "cost_set": cg.get("cost_set"), "cogs": cogs_total,
            "net": net,
            "net_pct": round(net / gross * 100, 1) if (net is not None and gross) else None,
            "match": cg.get("method", "нет"), "src": cg.get("source_sheet"),
        })
    view.sort(key=lambda x: (x["net"] is None, -(x["net"] or 0)))

    matched = [v for v in view if v["cogs"] is not None]
    unmatched = [v for v in view if v["cogs"] is None and v["qty"] > 0]

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "P&L + прибыль"
    cols = [
        ("sku", "SKU", 12), ("name", "Название", 40), ("qty", "Продано", 8),
        ("gross", "Выручка (с баллами)", 15),
        ("commission", "Комиссия", 11), ("logistics", "Логистика", 11),
        ("return_logistics", "Обр.лог.", 10), ("last_mile", "Посл.миля", 10),
        ("fbs_processing", "Обраб.FBS", 10), ("acquiring", "Эквайринг", 10),
        ("returns", "Возвраты", 11), ("contrib", "Вклад до COGS", 13),
        ("cost_set", "Себест/компл", 12), ("cogs", "Себест. всего", 13),
        ("net", "Чистая прибыль", 14), ("net_pct", "Марж.чист,%", 11),
        ("match", "Матч", 14), ("src", "Лист цен", 12),
    ]
    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True, size=10)
    for j, (k, t, w) in enumerate(cols, 1):
        c = ws.cell(1, j, t); c.fill = hf; c.font = hfont
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = w
    money = "#,##0;[Red]-#,##0"
    mcols = {"gross", "commission", "logistics", "return_logistics", "last_mile",
             "fbs_processing", "acquiring", "returns", "contrib", "cost_set", "cogs", "net"}
    for i, r in enumerate(view, 2):
        for j, (k, t, w) in enumerate(cols, 1):
            c = ws.cell(i, j, r.get(k))
            if k in mcols:
                c.number_format = money
            if k == "net" and r.get(k) is not None:
                c.font = Font(bold=True, color=("C00000" if r["net"] < 0 else "006100"))
            if k == "net_pct" and r.get(k) is not None:
                c.number_format = "0.0"
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Лист "Проверить себестоимость"
    ws2 = wb.create_sheet("Проверить себест-ть")
    for j, t in enumerate(["SKU", "Название", "Продано", "Выручка", "Вклад до COGS"], 1):
        ws2.cell(1, j, t).font = Font(bold=True)
    ws2.column_dimensions["B"].width = 46
    for i, r in enumerate(unmatched, 2):
        ws2.cell(i, 1, r["sku"]); ws2.cell(i, 2, r["name"])
        ws2.cell(i, 3, r["qty"]); ws2.cell(i, 4, r["gross"]).number_format = money
        ws2.cell(i, 5, r["contrib"]).number_format = money

    # Лист "Сводка магазина" — мост от выручки к чистой прибыли
    grand = data["grand_total_rub"]
    contrib_all = round(sum(v["contrib"] for v in view), 2)
    overhead = round(grand - contrib_all, 2)
    cogs_all = round(sum(v["cogs"] for v in matched), 2)
    store_net = round(grand - cogs_all, 2)
    ws3 = wb.create_sheet("Сводка магазина", 0)
    ws3.column_dimensions["A"].width = 46; ws3.column_dimensions["B"].width = 18
    bridge = [
        ("Показатель", "Сумма, ₽", True),
        ("Валовая выручка (с баллами за скидки)", round(sum(v["gross"] for v in view), 2), False),
        ("Расходы Ozon по SKU (комиссия, логистика, эквайринг, возвраты)",
         round(sum(sum(v[k] for k in SKU_COST_LINES) for v in view), 2), False),
        ("= Вклад товаров до COGS", contrib_all, True),
        ("Накладные магазина (реклама, штрафы — без SKU)", overhead, False),
        ("= Нетто по отчёту Ozon (до себестоимости)", grand, True),
        ("Себестоимость проданных товаров (COGS)", -cogs_all, False),
        ("= ЧИСТАЯ ПРИБЫЛЬ магазина", store_net, True),
    ]
    for i, (a, b, bold) in enumerate(bridge, 1):
        ca = ws3.cell(i, 1, a); cb = ws3.cell(i, 2, b)
        if bold:
            ca.font = Font(bold=True); cb.font = Font(bold=True)
        if isinstance(b, (int, float)):
            cb.number_format = "#,##0;[Red]-#,##0"
    ws3.cell(len(bridge) + 2, 1, "Себестоимость взята из файла цен, приоритет листа «Ozon FBS» (магазин преимущественно FBS).").font = Font(italic=True, size=9)

    wb.save(out_xlsx)
    net_total = round(sum(v["net"] for v in matched), 2)
    return {
        "out": out_xlsx, "sku_total": len(view),
        "matched": len(matched), "unmatched": len(unmatched),
        "gross": round(sum(v["gross"] for v in view), 2),
        "contrib_matched": round(sum(v["contrib"] for v in matched), 2),
        "cogs_matched": round(sum(v["cogs"] for v in matched), 2),
        "net_matched": net_total,
    }


if __name__ == "__main__":
    import sys, json
    print(json.dumps(build(sys.argv[1], sys.argv[2], sys.argv[3]), ensure_ascii=False, indent=2))
