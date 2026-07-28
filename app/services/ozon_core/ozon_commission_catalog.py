"""
marja.app — импорт ПОЛНОГО справочника комиссий Ozon (все категории).

Источник: официальный файл Ozon «Таблица категорий и типов для расчёта
вознаграждения» (seller-edu.ozon.ru -> Полный список комиссий -> кнопка
«Скачать таблицу категорий…»). Лист «Прайс РФ (БЗ)», ~9300 строк.

Структура (проверено 2026-07-27):
  строка 1 — группы схем (FBO / FBO Fresh / FBS / RFBS),
  строка 2 — заголовки: Основная категория | Категория | Тип товара |
             FBO(до100/100-300/свыше300) | FBO Fresh(...) | FBS(...) | RFBS(...),
  строки 3+ — данные.

Прогрессивная шкала: комиссия зависит от цены (до 100 / 100–300 / свыше 300 ₽).
"""
from __future__ import annotations
import json
import sys
import openpyxl

from . import ozon_unit

SHEET = "Прайс РФ (БЗ)"
# индексы колонок (0-based)
C = {"osn": 0, "cat": 1, "type": 2,
     "FBO": (3, 4, 5), "FBS": (9, 10, 11), "realFBS": (12, 13, 14)}
BANDS = ("le100", "le300", "gt300")  # до 100 / 100–300 / свыше 300


def _num(x):
    try:
        return round(float(x), 4)
    except (TypeError, ValueError):
        return None


def import_catalog(xlsx_path: str, sheet: str = SHEET) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    catalog: dict = {}
    n = 0
    for r in rows:
        if not r or len(r) <= C["type"]:
            continue
        cat = (str(r[C["cat"]]).strip() if r[C["cat"]] else "")
        typ = (str(r[C["type"]]).strip() if r[C["type"]] else "")
        if not cat and not typ:
            continue
        entry = {}
        for scheme, cols in (("FBO", C["FBO"]), ("FBS", C["FBS"]), ("realFBS", C["realFBS"])):
            vals = {}
            for band, ci in zip(BANDS, cols):
                v = _num(r[ci]) if ci < len(r) else None
                if v is not None:
                    vals[band] = v
            if vals:
                entry[scheme] = vals
        if entry:
            catalog.setdefault(cat, {})[typ] = entry
            n += 1
    # дефолт по категории = модальный тип (по FBS gt300)
    for cat, types in catalog.items():
        from collections import Counter
        key = Counter(
            (t.get("FBO", {}).get("gt300"), t.get("FBS", {}).get("gt300"), t.get("realFBS", {}).get("gt300"))
            for t in types.values()).most_common(1)[0][0]
        types["_default"] = {
            "FBO": {"gt300": key[0]} if key[0] is not None else {},
            "FBS": {"gt300": key[1]} if key[1] is not None else {},
            "realFBS": {"gt300": key[2]} if key[2] is not None else {},
        }

    s = ozon_unit.load_settings()
    s["commission_catalog"] = catalog
    s.setdefault("category_aliases", {})
    s["commission_catalog_meta"] = {"source_file": xlsx_path.split("/")[-1], "categories": len(catalog), "types": n}
    with open(ozon_unit.SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(s, f, ensure_ascii=False, indent=2)
    return {"categories": len(catalog), "types": n}


if __name__ == "__main__":
    print(json.dumps(import_catalog(sys.argv[1]), ensure_ascii=False, indent=2))
