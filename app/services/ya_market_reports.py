"""Yandex.Market Reports API — асинхронные отчёты.

Общая схема работы с отчётом:
  1. POST /businesses/{businessId}/reports/{report_type}/generate     → {reportId}
  2. GET  /reports/info/{reportId}  (полить каждые 5 сек)             → status: PENDING|PROCESSING|DONE|FAILED
  3. Когда DONE — в ответе есть file (прямой URL на XLSX)
  4. Скачиваем XLSX и парсим.

Мы используем:
  - key-indicators           → показатели бизнеса (быстро, для сводки)
  - goods-realization        → реестр реализации (детально по SKU за период)
  - united-marketplace-services → услуги маркетплейса (комиссии/логистика/хранение по периодам)
  - united-orders            → единый реестр заказов (для точного qty по SKU)
"""
from __future__ import annotations
import time
import io
from datetime import date, datetime, timedelta
from typing import Any, Iterable
import httpx
from openpyxl import load_workbook

BASE_URL = "https://api.partner.market.yandex.ru"

MAX_WAIT_SEC = 180
POLL_INTERVAL_SEC = 5


class YaMarketReportError(Exception):
    """Ошибка при работе с отчётами Я.Маркета."""


class YaMarketReportsClient:
    """Клиент для генерации и скачивания асинхронных отчётов Ya.Market."""

    def __init__(self, api_token: str, business_id: int, timeout: float = 30.0):
        self.token = api_token
        self.business_id = business_id
        self._client = httpx.Client(
            base_url=BASE_URL,
            headers={
                "Authorization": f"Bearer {api_token}",
                "Api-Key": api_token,
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            timeout=timeout,
        )

    def __enter__(self): return self
    def __exit__(self, *a): self._client.close()

    def close(self): self._client.close()

    def _post_generate(self, report_type: str, body: dict) -> str:
        url = f"/businesses/{self.business_id}/reports/{report_type}/generate"
        r = self._client.post(url, params={"format": "FILE"}, json=body)
        if r.status_code >= 400:
            raise YaMarketReportError(
                f"POST {url} → {r.status_code}: {r.text[:400]}"
            )
        return r.json().get("result", {}).get("reportId", "")

    def generate_key_indicators(self, dt_from: date, dt_to: date) -> str:
        return self._post_generate("key-indicators", {
            "businessId": self.business_id,
            "dateFrom": dt_from.isoformat(),
            "dateTo": dt_to.isoformat(),
            "reportPeriod": "WEEK",
        })

    def generate_realization(self, dt_from: date, dt_to: date, campaign_id: int | None = None) -> str:
        body = {
            "businessId": self.business_id,
            "dateFrom": dt_from.isoformat(),
            "dateTo": dt_to.isoformat(),
        }
        if campaign_id:
            body["campaignId"] = campaign_id
        return self._post_generate("goods-realization", body)

    def generate_marketplace_services(self, dt_from: date, dt_to: date) -> str:
        return self._post_generate("united-marketplace-services", {
            "businessId": self.business_id,
            "dateTimeFrom": dt_from.isoformat() + "T00:00:00Z",
            "dateTimeTo": dt_to.isoformat() + "T23:59:59Z",
        })

    def generate_united_orders(self, dt_from: date, dt_to: date) -> str:
        return self._post_generate("united-orders", {
            "businessId": self.business_id,
            "dateFrom": dt_from.isoformat(),
            "dateTo": dt_to.isoformat(),
        })

    def get_report_info(self, report_id: str) -> dict:
        r = self._client.get(f"/reports/info/{report_id}")
        if r.status_code >= 400:
            raise YaMarketReportError(
                f"GET /reports/info/{report_id} → {r.status_code}: {r.text[:400]}"
            )
        return r.json().get("result", {})

    def wait_for_report(self, report_id: str,
                         max_wait_sec: int = MAX_WAIT_SEC,
                         poll_interval_sec: int = POLL_INTERVAL_SEC) -> dict:
        started = time.monotonic()
        while True:
            info = self.get_report_info(report_id)
            status = (info.get("status") or "").upper()
            if status == "DONE":
                return info
            if status == "FAILED":
                raise YaMarketReportError(
                    f"Report {report_id} FAILED: {info.get('subStatus') or info}"
                )
            if time.monotonic() - started > max_wait_sec:
                raise YaMarketReportError(
                    f"Report {report_id} timeout after {max_wait_sec}s (status={status})"
                )
            time.sleep(poll_interval_sec)

    def download_report(self, file_url: str) -> bytes:
        with httpx.Client(timeout=60.0) as cli:
            r = cli.get(file_url, follow_redirects=True)
            r.raise_for_status()
            return r.content


def _find_row_by_prefix(ws, prefix: str, col: int = 1, start: int = 1, end: int | None = None) -> int | None:
    p = prefix.lower().strip()
    end = end or ws.max_row
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=col).value
        if v and str(v).lower().strip().startswith(p):
            return r
    return None


def _num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".").replace("₽", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_key_indicators_xlsx(xlsx_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    labels = {
        "revenue":    ["стоимость реализованного", "выручка"],
        "commission": ["комиссия"],
        "logistics":  ["логистик", "услуги по доставк"],
        "storage":    ["хранени"],
        "ads":        ["продвиж", "реклам"],
        "penalties":  ["штраф"],
        "acquiring":  ["эквайринг", "прием оплаты"],
        "units":      ["количество", "проданных единиц", "штук"],
    }

    def _values_from_row(row_idx: int) -> list[float]:
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        return [_num(v) for v in row[1:] if v is not None and v != ""]

    result = {"totals": {}, "weeks": []}
    for metric, prefixes in labels.items():
        row_idx = None
        for p in prefixes:
            row_idx = _find_row_by_prefix(ws, p)
            if row_idx:
                break
        if not row_idx:
            result["totals"][metric] = 0.0
            continue
        vals = _values_from_row(row_idx)
        result["totals"][metric] = vals[0] if vals else 0.0
        result.setdefault("_rows", {})[metric] = vals

    return result


def parse_realization_xlsx(xlsx_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(20, ws.max_row) + 1):
        row = [str(c.value or "").lower() for c in ws[r]]
        joined = " ".join(row)
        if any(k in joined for k in ["sku продавца", "ваш sku", "артикул продавца", "shop sku"]):
            header_row = r
            break
    if not header_row:
        return []

    header_map: dict[str, int] = {}
    for i, cell in enumerate(ws[header_row], start=1):
        v = str(cell.value or "").strip().lower()
        if not v:
            continue
        if "ваш sku" in v or "sku продавца" in v or "артикул продавца" in v or "shop sku" in v:
            header_map["sku"] = i
        elif "название" in v or "наименование" in v:
            header_map["name"] = i
        elif "количество" in v:
            header_map["qty"] = i
        elif "цена продажи" in v or "цена товара" in v or "цена реализации" in v:
            header_map["price"] = i
        elif "комиссия" in v:
            header_map["commission"] = i
        elif "доставк" in v or "логистик" in v:
            header_map["delivery"] = i
        elif "эквайринг" in v:
            header_map["acquiring"] = i

    items: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        sku = str(ws.cell(row=r, column=header_map.get("sku", 0)).value or "").strip() if header_map.get("sku") else ""
        if not sku:
            continue
        item = {"sku": sku}
        for key in ("name",):
            col = header_map.get(key)
            item[key] = str(ws.cell(row=r, column=col).value or "").strip() if col else ""
        for key in ("qty", "price", "commission", "delivery", "acquiring"):
            col = header_map.get(key)
            item[key] = _num(ws.cell(row=r, column=col).value) if col else 0.0
        items.append(item)
    return items


def week_ranges(dt_from: date, dt_to: date) -> list[tuple[date, date]]:
    weeks: list[tuple[date, date]] = []
    cur = dt_from
    while cur <= dt_to:
        wk_start = cur - timedelta(days=cur.weekday())
        wk_end = wk_start + timedelta(days=6)
        if wk_start < dt_from:
            wk_start = dt_from
        if wk_end > dt_to:
            wk_end = dt_to
        weeks.append((wk_start, wk_end))
        cur = wk_end + timedelta(days=1)
    return weeks

